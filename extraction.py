# -*- coding: utf-8 -*-
"""
Created on Mon Jul  8 10:12:45 2019

@author: gasha
"""

import cv2 
import numpy as np
from path import Path as path
import math
import skimage
from openpyxl import Workbook
#from sklearn.preprocessing import MinMaxScaler


#path to directory holding images
paths = []
paths.append(path('/Users/gasha/.spyder-py3/bhw1'))
paths.append(path('/Users/gasha/.spyder-py3/bhl1'))
#paths.append(path('/Users/gasha/.spyder-py3/shoreline'))
#paths.append(path('/Users/gasha/.spyder-py3/water'))
#paths.append(path('/Users/gasha/.spyder-py3/land'))

#variable initialization
laparr=[]
canarr=[]
harr=[]
barr=[]
garr=[]
rarr=[]
grayarr=[]
ls=[]
lm=[]
ld=[]
cs=[]
cm=[]
cd=[]
hm=[]
bm=[]
gm=[]
rm=[]
e=[]
fc=[]
count=0
fcount= 0

#main loop
for p in paths:
    for f in p.files('*.jpg'):
        img = cv2.imread(f)
        h,w,_ = img.shape
        
        #creates images for extraction
        gray= cv2.cvtColor(img,cv2.COLOR_BGR2GRAY)
        lap = cv2.Laplacian(gray, cv2.CV_8U, ksize=15)
        hsv= cv2.cvtColor(img,cv2.COLOR_BGR2HSV)
        can=cv2.Canny(img,100,200)
        
        #adds wanted pixels to arrays
        for k in range(h):
            for j in range(w):
                if img[k,j,2] >= 250 and img[k,j,0] >= 250 and img[k,j,1] == 0:
                    continue
                else:
                    laparr.append(lap[k,j])
                    canarr.append(can[k,j])
                    harr.append(hsv[k,j,0])
                    barr.append(img[k,j,0])
                    garr.append(img[k,j,1])
                    rarr.append(img[k,j,2])
                    grayarr.append(gray[k,j])
                    
        #preforms calculations to get the data
        lapsum = sum(laparr)
        lapmean = lapsum/len(laparr)
        lapdev = math.sqrt(np.var(laparr))
        cansum = sum(canarr)
        canmean = cansum/len(canarr)
        candev = math.sqrt(np.var(canarr))
        hmean = sum(harr)/len(harr)
        bmean = sum(barr)/len(barr)
        gmean = sum(garr)/len(garr)
        rmean = sum(rarr)/len(rarr)
        
        #entropy calculation
        entropy = skimage.measure.shannon_entropy(grayarr)
        
        print(count)
        
        #adds data for averaging
        #ls.append(lapsum)
        lm.append(lapmean)
        ld.append(lapdev)
        #cs.append(cansum)
        cm.append(canmean)
        cd.append(candev)
        hm.append(hmean)
        bm.append(bmean)
        gm.append(gmean)
        rm.append(rmean)
        e.append(entropy)
        fc.append(fcount)
        

        #emptys image arrays
        laparr.clear()
        canarr.clear()
        harr.clear()
        barr.clear()
        garr.clear()
        rarr.clear()
        grayarr.clear()
        
        #counter
        count += 1
    
    #used to label classes
    fcount += 1
    

#xlsx intiialization
wb = Workbook()
ws = wb.create_sheet("shoreline")

#labels
#c=ws.cell(row=1,column=1)
#c.value = "laplacian Sum"
c=ws.cell(row=1,column=1)
c.value = "Laplacian Mean"
c=ws.cell(row=1,column=2)
c.value = "Laplacian Standard Deviation"
#c=ws.cell(row=1,column=4)
#c.value = "Canny Sum"
c=ws.cell(row=1,column=3)
c.value = "Canny Mean"
c=ws.cell(row=1,column=4)
c.value = "Canny Standard Deviation"
c=ws.cell(row=1,column=5)
c.value = "Hue Mean"
c=ws.cell(row=1,column=6)
c.value = "Blue Mean"
c=ws.cell(row=1,column=7)
c.value = "Green Mean"
c=ws.cell(row=1,column=8)
c.value = "Red Mean"
c=ws.cell(row=1,column=9)
c.value = "Shannons Entropy"
c=ws.cell(row=1,column=10)
c.value = "Class"

#saves data
n = 0
for n in range(count):

    c=ws.cell(row=n+2,column=1)
    c.value = lm[n]
    c=ws.cell(row=n+2,column=2)
    c.value = ld[n]

    c=ws.cell(row=n+2,column=3)
    c.value = cm[n]
    c=ws.cell(row=n+2,column=4)
    c.value = cd[n]
    c=ws.cell(row=n+2,column=5)
    c.value = hm[n]
    c=ws.cell(row=n+2,column=6)
    c.value = bm[n]
    c=ws.cell(row=n+2,column=7)
    c.value = gm[n]
    c=ws.cell(row=n+2,column=8)
    c.value = rm[n]
    c=ws.cell(row=n+2,column=9)
    c.value = e[n]
    c=ws.cell(row=n+2,column=10)
    c.value = fc[n]
    
    #saves xlsx document
    wb.save('bhv1.xlsx')
